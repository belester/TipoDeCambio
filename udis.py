from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from seleniumbase import Driver
from tabulate import tabulate
from datetime import datetime
from datetime import datetime, timedelta
from pathlib import Path
import pandas as pd
import openpyxl
import calendar
import pyodbc
import time
from os import remove
from os import path

def UDIS_P():

    #print(fecha_Inicio)
    #print(fecha_Actual)

    server = '15.10.155.82\DWHDES001'
    database = 'Rentabilidad'
    #----------------------
    try:
        # Conexión a SQL Server
        conn = pyodbc.connect(
            'Driver={SQL Server};Server=' + server + ';Database=' + database + ';Trusted_Connection=yes;')
        print('Conexión Exitosa')

    except Exception as e:
        print(f'Error al intentar conectarse debido a: {e}')

    # Consulta a la base de datos
    cursor = conn.cursor()

    cursor.execute('SELECT MAX(Fecha) FROM C0001 WHERE id_moneda=3;')
    ult_fechaBD = cursor.fetchone()[0]

    # Prueba de impresión, recordar comentar
    # print(f'Ultima Fecha obtenida por la BD: {ult_fechaBD} \n')

    # Cerrar la consulta
    cursor.close()

    # Obtener la fecha actual
    fecha_actual = datetime.now().date()

    # Función para obtener el número de días de un mes
    def obtener_dias_mes(mes, anio):
        if mes == 2:  # Febrero
            if calendar.isleap(anio):  # Verificar si el año es bisiesto
                return 29
            else:
                return 28
        else:
            return calendar.monthrange(anio, mes)[1]

    # Convertir la fecha a un objeto datetime, si no es None
    if ult_fechaBD:
        if isinstance(ult_fechaBD, str):
            ult_fechaBD = datetime.strptime(ult_fechaBD, "%Y-%m-%d").date()
        elif isinstance(ult_fechaBD, datetime.datetime):
            ult_fechaBD = ult_fechaBD.date()

    # Recordar comentar la siguiente linea de código
    # print(f'Fecha después de la conversión: {ult_fechaBD}')

    # Verificar si se encontró una fecha en la base de datos
    if ult_fechaBD:
        # Obtener el siguiente día de la última fecha en la base de datos
        fecha_inicio = ult_fechaBD + timedelta(days=1)
        print(fecha_inicio)
    else:
        # No se encontró una fecha en la base de datos, comenzar desde el primer día del mes actual
        fecha_inicio = fecha_actual.replace(day=1).date()
    #----------------------
    driver = Driver(uc=True)

    UDIS = "https://www.banxico.org.mx/SieInternet/consultarDirectorioInternetAction.do?accion=consultarCuadro&idCuadro=CP150&locale=es"

    driver.get(UDIS)
    # Esperar hasta que el botón esté visible
    wait = WebDriverWait(driver, 30)

    button = wait.until(EC.visibility_of_element_located((By.ID, "exportaCuadroToggle")))

    # Hacer clic en el botón
    button.click()

    # Intenta esperar un poco para asegurarte de que la página esté completamente cargada
    driver.implicitly_wait(10)

    # Cuadro Fecha Inicial
    fecha_inicio_input = wait.until(EC.visibility_of_element_located((By.ID, "expCuadroFechaInicio")))

    # Cuadro Final
    fecha_fin_input = wait.until(EC.visibility_of_element_located((By.ID, "expCuadroFechaFinal")))

    # Borrar la información existente en los campos de entrada
    fecha_inicio_input.clear()
    fecha_fin_input.clear()
    Fecha_Actual = fecha_actual + timedelta(days=7)
    # Escribir la fecha final en el campo de entrada
    fecha_fin_input.send_keys(Fecha_Actual.strftime("%d-%m-%Y"))

    # Escribir la fecha inicial en el campo de entrada
    fecha_inicio_input.send_keys(fecha_inicio.strftime("%d-%m-%Y"))

    #DAR ENTER
    driver.switch_to.active_element.send_keys(Keys.ENTER)

    #Conexion a la bd
    try:
        # Conexión a SQL Server
        conn = pyodbc.connect(
            'Driver={SQL Server};Server=' + server + ';Database=' + database + ';Trusted_Connection=yes;')
        print('Conexión Exitosa')

    except Exception as e:
        print(f'Error al intentar conectarse debido a: {e}')

    #Escoge el archivo mas reciente
    dir_actual = Path.cwd()
    files_csv = dir_actual.glob('downloaded_files/*.xlsx')
    print(files_csv)
    latest_file = max(files_csv)
    print(latest_file)

    excel_dataframe = openpyxl.load_workbook(latest_file)

    dataframe = excel_dataframe.active

    print(dataframe)

    data = []
    for row in range(9, dataframe.max_row):
        _row = [row]
        for col in dataframe.iter_cols(3, dataframe.max_column):
            _row.append(col[row].value)

        data.append(_row)

    fecha = data[0]
    udis = data[2]

    #Eliminamos el primer elemento de la lista
    fecha.pop(0)
    udis.pop(0)


    cursor = conn.cursor()
    # Insertar registros
    for fecha, udis in zip(fecha, udis):
        cursor.execute("INSERT INTO C0001 VALUES(?,3,?,null)",
                       datetime.strptime(fecha, "%d/%m/%Y"), udis)


    # Elimina los archivos excel
    if path.exists(latest_file):
        remove(latest_file)

    conn.commit()
    time.sleep(6)
    driver.quit()