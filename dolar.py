from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from pathlib import Path
from os import remove
from os import path
import pyexcel as p
import pandas as pd
import openpyxl
import pyodbc
import time
from selenium import webdriver
from conChrome import conCromeDriver
from ConexionBD.Con82 import conServidor


def dolar(fecha_Actual, fecha_Inicio):
    driver = conCromeDriver()
   # print(fecha_Inicio)
   # print(fecha_Actual)

    UDIS = "https://www.banxico.org.mx/tipcamb/tipCamMIAction.do?idioma=sp"
    driver.get(UDIS)

    # Esperar hasta que el botón esté visible
    wait = WebDriverWait(driver, 30)

    # Intenta esperar un poco para asegurarte de que la página esté completamente cargada
    driver.implicitly_wait(10)

    # Cuadro Fecha Inicial
    fecha_inicio_input = wait.until(EC.visibility_of_element_located((By.NAME, "fechaInicial")))

    # Cuadro Final
    fecha_fin_input = wait.until(EC.visibility_of_element_located((By.NAME, "fechaFinal")))

    formato_xls = wait.until(EC.visibility_of_element_located((By.NAME, "salida")))

    # Borrar la información existente en los campos de entrada
    fecha_inicio_input.clear()
    fecha_fin_input.clear()

    # Escribir la fecha final en el campo de entrada
    fecha_fin_input.send_keys(fecha_Actual.strftime("%d-%m-%Y"))

    # Escribir la fecha inicial en el campo de entrada
    fecha_inicio_input.send_keys(fecha_Inicio.strftime("%d-%m-%Y"))

    #Seleccionar el formato
    formato_xls.send_keys("xls")

    #DAR ENTER
    driver.switch_to.active_element.send_keys(Keys.ENTER)

    # click boton
    button = wait.until(EC.visibility_of_element_located((By.CLASS_NAME, "botonesSIE")))

    #click boton
    button.click()

    cursor = conServidor()

#----------------------------------------------------------------------------------
   #cambiar de formato el EXCEL
    p.save_book_as(file_name='downloaded_files/tipoCambio.xls',
                   dest_file_name='downloaded_files/tipoCambio.xlsx')

    #Buscar el formato mas reciente
    dir_actual = Path.cwd()
    files_csv = dir_actual.glob('downloaded_files/*.xlsx')
    latest_file = max(files_csv)
    print(latest_file)

    excel_dataframe = openpyxl.load_workbook(latest_file)

    dataframe = excel_dataframe.active

    print(dataframe)

    data = []
    for row in range(8, dataframe.max_row):
        _row = [row]
        for col in dataframe.iter_cols(1, dataframe.max_column):
            _row.append(col[row].value)

        data.append(_row)

    df = pd.DataFrame(data)
  #  df.drop(df.columns[[0, 2, 3]], axis=1, inplace=True)

    fecha = df[1]
    tipoDeCambioDolar = df[2]

    dolaArray = []
    for elemtos in tipoDeCambioDolar:
        dolaArray.append(elemtos)

    for i in range(len(dolaArray)):
        if dolaArray[i] == 'N/E':
            if i > 0:
                dolaArray[i] = dolaArray[i - 1]

#---------------------------------------------------------------------------

    cursor = conn.conServidor()
    # Insertar registros
    for fecha, dolaArray in zip(fecha, dolaArray):
        try:
          cursor.execute("INSERT INTO C0001 VALUES(?,2,?,null)", (fecha, dolaArray))
          conn.commit()
        except pyodbc.IntegrityError:
            print("Un registro con la misma fecha e id_moneda ya existe.")
            cursor.execute("UPDATE C0001 SET Valor = ? WHERE Fecha = ? AND id_moneda = 2", (dolaArray, fecha))
            conn.commit()

    #Elimina los archivos excel
    if path.exists("downloaded_files/tipoCambio.xls"):
        remove('downloaded_files/tipoCambio.xls')

    # Elimina los archivos excel
    if path.exists("downloaded_files/tipoCambio.xlsx"):
        remove('downloaded_files/tipoCambio.xlsx')

    conn.commit()
    time.sleep(6)
    driver.quit()